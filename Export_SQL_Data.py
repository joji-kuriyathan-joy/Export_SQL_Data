##############################################################
'''
Export_SQL_Data_V1.0                           June-10-2020
This python file executes SQL statements and Export Data into EXCEL/CSV
Options to Export with header and with out headers
Execution process is written down to Log file in Log Folder
Exported data in the form of EXCEL/CSV is present in the Output Folder after Execution
If the execution type is EXCEL , then the  single Excel contains sheets of executed query from the query_list
Note:
    Make sure that the sheet_list and csv_filename_list is equal to the query_list in the config.
    Specify the Type of Export in the config
    Specify the Output and Log Folder path in the config
    Give SQL connection details in the config. The DB used is SQL Server
    Change to True / False for with column headers and with out
    Change the delimiter to ',' or '|' while exporting to CSV
'''
###############################################################

import json
import os
import pyodbc
import pandas as pd
import datetime
import sys
import traceback
from openpyxl import load_workbook
from openpyxl import Workbook

# config_file
config_file = "SQL_Config.json"


def create_folder(folderPath):
    if not os.path.exists(folderPath):
        os.makedirs(folderPath)


def getConnection(db_host, db_name, db_username, db_password, db_port, log_file):
    try:
        print("_________ Connecting to SQL Server: %s" % db_host)
        log_file.write("\n_________ Connecting to SQL Server: %s" % db_host)
        if db_port == "":
            conn = pyodbc.connect(
                'DRIVER={SQL Server};SERVER=' + str(db_host) + ';DATABASE=' + str(db_name) + ';UID=' + str(
                    db_username) + ';PWD=' + str(db_password) + '')
        else:
            # conn = pyodbc.connect("Driver={FreeTDS};Server=192.168.2.96;uid=beta\\mmuthusamy;pwd=Cairocairo4551_;database=MM_EH001_20200212;port=1433")
            conn = pyodbc.connect(
                'Driver={FreeTDS};Server=' + str(servername) + ';database=' + str(dbname) + ';uid=' + str(uname) + ';pwd=' + str(pswd) + ';port=' + str(port))

        print("_________ Connection Sucessfull for : %s." % db_host)
        log_file.write("\n_________ Connection Sucessfull.")
        return conn


    except Exception as e:
        print("### Error in getConnection (): %s" % e)
        print("Traceback: ", traceback.format_exc())
        log_file.write("\n ### Error in SQL Connection : %s  \n Traceback : %s" % (e, traceback.format_exc()))


def Export_to_Excel(connection, query_list, sheets_list, column_header_flag, output_file_name, output_file_path,
                    log_file):
    ''' Export data to excel'''
    try:
        curr_date_str = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
        output_file_name = output_file_name.replace(
            "$$YYYY-MM-DD-HH-MM-SS$$", curr_date_str)
        out_excel_file_name = os.path.join(output_file_path, output_file_name + ".xlsx")
        if len(query_list) == len(sheets_list):
            if os.path.isfile(out_excel_file_name):
                workbook = load_workbook(out_excel_file_name)
            else:
                workbook = Workbook()
            print("Output Excel file name : ", out_excel_file_name)
            log_file.write("\n Output Excel file name : %s" % out_excel_file_name)
            print("Include Column Headers : ", column_header_flag)
            log_file.write("\n Include Column Headers : %s" % column_header_flag)
            for i, query in enumerate(query_list):
                if not os.path.isfile(out_excel_file_name) and i == 0:
                    worksheet = workbook.active
                    worksheet.title = sheets_list[i]
                else:
                    worksheet = workbook.create_sheet(sheets_list[i])

                log_file.write(
                    "\n Query Execution Start Date: " + datetime.datetime.now().strftime("%Y-%m-%d-%H-%M-%S"))
                log_file.write("\n --------------- Query --------------- \n" + query)
                print("\n--------------- Query --------------- \n" + query)
                print("Query Execution Start Date: " + datetime.datetime.now().strftime("%Y-%m-%d-%H-%M-%S"))

                cursor = connection.cursor()
                cursor.execute(query)
                records = cursor.fetchall()

                print("Query Execution End Date: " + datetime.datetime.now().strftime("%Y-%m-%d-%H-%M-%S"))
                log_file.write("\n Query Execution End Date: " + datetime.datetime.now().strftime("%Y-%m-%d-%H-%M-%S"))

                columns = [desc[0] for desc in cursor.description]
                log_file.write("\n Populating data Start Date: " + datetime.datetime.now().strftime("%Y-%m-%d-%H-%M-%S"))
                if column_header_flag == "True":
                    for n, column in enumerate(columns, 0):
                        worksheet.cell(row=1, column=n + 1).value = column

                        for n, row in enumerate(records, 1):
                            for i in range(len(columns)):
                                worksheet.cell(row=n + 1, column=i + 1).value = row[i]
                else:
                    for n, row in enumerate(records, 1):
                        for i in range(len(columns)):
                            worksheet.cell(row=n, column=i + 1).value = row[i]
                log_file.write(
                    "\n Populating data End Date: " + datetime.datetime.now().strftime("%Y-%m-%d-%H-%M-%S"))
            workbook.save(out_excel_file_name)
            print("Excel file created : ",out_excel_file_name)
            log_file.write("\n_________ Excel file Created with sheets : [%s]"%out_excel_file_name)
        else:
            print(
                "The number of sheet list is not matching with the query list. Make the both list equal.  Excel not created.")
            log_file.write(
                "\n The number of sheet list is not matching with the query list. Make the both list equal. Excel not created.")
    except Exception as e:
        print("### Error in Export_to_Excel (): %s" % e)
        print("Traceback: ", traceback.format_exc())
        log_file.write("\n ### Error in Export_to_Excel : %s  \n Traceback : %s" % (e, traceback.format_exc()))


def Export_to_CSV(connection, query_list, csv_file_name_list, delimeter, output_file_path, column_header_flag,
                  log_file):
    '''Export data to csv with specified delimeter'''
    try:
        print("Output Folder path : ", output_file_path)
        log_file.write("\n Output Folder path : %s" % output_file_path)
        print("Include Column Headers : ", column_header_flag)
        log_file.write("\n Include Column Headers : %s" % column_header_flag)
        if len(query_list) == len(csv_file_name_list):
            for i, query in enumerate(query_list):
                log_file.write(
                    "\n Query Execution Start Date: " + datetime.datetime.now().strftime("%Y-%m-%d-%H-%M-%S"))
                log_file.write("\n --------------- Query --------------- \n" + query)
                print("\n--------------- Query --------------- \n" + query)
                print("Query Execution Start Date: " + datetime.datetime.now().strftime("%Y-%m-%d-%H-%M-%S"))

                SQL_Query = pd.read_sql_query(query, connection)
                df = pd.DataFrame(SQL_Query)

                print("Query Execution End Date: " + datetime.datetime.now().strftime("%Y-%m-%d-%H-%M-%S"))
                log_file.write("\n Query Execution End Date: " + datetime.datetime.now().strftime("%Y-%m-%d-%H-%M-%S"))

                csv_file_name = os.path.join(output_file_path, csv_file_name_list[i] + "_" + str(
                    datetime.datetime.now().strftime("%Y%m%d-%H%M%S")) + ".csv")
                log_file.write(
                    "\n Populating data Start Date: " + datetime.datetime.now().strftime("%Y-%m-%d-%H-%M-%S"))

                if column_header_flag == "True":
                    df.to_csv(csv_file_name, sep=delimeter, index=False)
                else:
                    df.to_csv(csv_file_name, sep=delimeter, index=False, header=False)
                log_file.write(
                    "\n Populating data End Date: " + datetime.datetime.now().strftime("%Y-%m-%d-%H-%M-%S"))
                print("File Created : ", csv_file_name)
                log_file.write("\n File created : %s" % csv_file_name)

        else:
            print(
                "The number of csv file name  list is not matching with the query list. Make the both list equal.  Excel not created.")
            log_file.write(
                "\n The number of csv file name list is not matching with the query list. Make the both list equal. Excel not created.")


    except Exception as e:
        print("### Error in Export_to_CSV (): %s" % e)
        print("Traceback: ", traceback.format_exc())
        log_file.write("\n ### Error in Export_to_CSV : %s  \n Traceback : %s" % (e, traceback.format_exc()))


def SQL_executor(config_file):
    try:
        print("====================== Export SQL Data ======================")

        try:
            with open(config_file, 'r') as f:
                datastore = json.load(f)
        except Exception as e:
            sys.exit(
                "--------- Failed to Load Data from Config file (%s). File may not exist or it is empty." % config_file)

        logFilePath = datastore["General"]["Log_File_Path"]
        create_folder(logFilePath)
        curr_date = datetime.datetime.now()
        curr_date_str = curr_date.strftime("%Y%m%d-%H%M%S")

        print("Creating New Log File...")
        log_file = open(logFilePath + "Log_" + curr_date_str + ".txt", "w+")
        print("Created New Log File: ", log_file)

        log_file.write(
            "\n====================== Export SQL Data ======================[%s]" % str(datetime.datetime.now()))
        if bool(datastore):
            print("_________Config File read successfully (%s)." % config_file)
            log_file.write("\n_________ Config File read successfully (%s)." % config_file)
        else:
            log_file.write("\n_________ Failed to read Config File : [%s]" % config_file)
            sys.exit("Failed to read Config File (%s)." % config_file)

        output_file_path = datastore["General"]["Out_File_Path"]
        output_file_name = datastore["General"]["Output_Excel_File_Name"]

        print("_________ Getting Connection for SQL ")
        log_file.write("\n_________ Getting Connection for SQL ")

        db_host = datastore["SQL_Details"]["Connection_Details"]["DBHost"]
        db_name = datastore["SQL_Details"]["Connection_Details"]["DB_Name"]
        db_username = datastore["SQL_Details"]["Connection_Details"]["DBUserName"]
        db_password = datastore["SQL_Details"]["Connection_Details"]["DBPassword"]
        db_port = datastore["SQL_Details"]["Connection_Details"]["DBPort"]

        connection = getConnection(db_host, db_name, db_username, db_password, db_port, log_file)

        query_list = datastore["SQL_Details"]["Query_List"]
        sheet_list = datastore["SQL_Details"]["Sheet_list"]
        export_type = datastore["SQL_Details"]["Export_Type"]
        column_header_flag = datastore["SQL_Details"]["Column_Headers"]
        delimeter = datastore["SQL_Details"]["Delimiter"]
        csv_file_name_list = datastore["SQL_Details"]["CSV_FileName_List"]
        if export_type.lower().strip() == "excel":
            print("_________ Export Type Excel ")
            print("Start time: %s" % datetime.datetime.now().strftime("%Y%m%d-%H%M%S"))
            log_file.write("\n_________ Export Type : [Excel] ")
            log_file.write("\n Start time: %s" % str(datetime.datetime.now().strftime("%Y%m%d-%H%M%S")))
            Export_to_Excel(connection, query_list, sheet_list, column_header_flag, output_file_name, output_file_path,
                            log_file)
            print("End time: %s" % datetime.datetime.now().strftime("%Y%m%d-%H%M%S"))
            log_file.write("\n End time: %s" % str(datetime.datetime.now().strftime("%Y%m%d-%H%M%S")))

        elif export_type.lower().strip() == "csv":
            print("_________ Export Type CSV ")
            log_file.write("\n_________ Export Type : [CSV] ")
            print("Start time: %s" % datetime.datetime.now().strftime("%Y%m%d-%H%M%S"))
            Export_to_CSV(connection, query_list, csv_file_name_list, delimeter, output_file_path, column_header_flag,
                          log_file)
        else:
            if export_type.lower().strip() == "":
                print("_________ Export Type is Empty ")
                log_file.write("\n_________ Export Type is Empty. Specify the type as EXCEL / CSV ")

        log_file.write(
            "\n====================== END ======================[%s]" % str(datetime.datetime.now()))




    except Exception as e:
        print("### Error in SQL Executor () : %s" % e)
        print("Traceback : ", traceback.format_exc())


SQL_executor(config_file)
